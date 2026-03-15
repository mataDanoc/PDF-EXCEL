"""
excel_renderer.py  -  Visual-fidelity Excel generation

Creates an .xlsx file that visually matches the original PDF layout.

Key techniques:
  - Grid lines HIDDEN so the sheet looks like a clean document, not a grid
  - Dynamic column widths matched to PDF coordinate distances
  - Cell merging for text elements that span multiple boundary columns
  - Background colours extracted from PDF filled rectangles
  - Font colours (white on dark backgrounds, black otherwise)
  - Bold and font-size preserved from PDF metadata
  - Thin borders around table regions
  - Row heights proportional to PDF line spacing
  - Alignment: numbers right-aligned, headers centred, text left-aligned
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Dict, List, Set, Tuple

from .config import Config, DEFAULT_CONFIG
from .grid_builder import ExcelGrid, GridCell

logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side,
    )
    from openpyxl.utils import get_column_letter
    _OK = True
except ImportError:
    _OK = False
    logger.error("openpyxl not installed.  pip install openpyxl")

# PDF pt -> Excel character-width conversion
_PT_TO_CHAR = 1.0 / 5.25

# Pre-built styles (created lazily)
_THIN_SIDE = None
_THIN_BORDER = None


def _ensure_styles():
    global _THIN_SIDE, _THIN_BORDER
    if _THIN_SIDE is None:
        _THIN_SIDE = Side(border_style="thin", color="FF999999")
        _THIN_BORDER = Border(
            left=_THIN_SIDE, right=_THIN_SIDE,
            top=_THIN_SIDE, bottom=_THIN_SIDE,
        )


class ExcelRenderer:
    def __init__(self, config: Config = DEFAULT_CONFIG) -> None:
        self.cfg = config

    def render(self, grids: List[ExcelGrid], output_path: str | Path) -> Path:
        if not _OK:
            raise RuntimeError("openpyxl required. pip install openpyxl")

        out = Path(output_path)
        out.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        wb.remove(wb.active)

        for grid in grids:
            self._render_page(wb, grid)

        if not wb.worksheets:
            wb.create_sheet("Empty")

        wb.save(str(out))
        logger.info("Saved Excel file: %s", out)
        return out

    # ── Per-page rendering ────────────────────────────────────────────────

    def _render_page(self, wb: "Workbook", grid: ExcelGrid) -> None:
        _ensure_styles()
        cfg = self.cfg

        ws = wb.create_sheet(title=f"Page {grid.page_num}")

        # ─────────── HIDE GRID LINES ─────────────────────────────────────
        ws.sheet_view.showGridLines = False

        # ─────────── COLUMN WIDTHS ───────────────────────────────────────
        for i, w_pt in enumerate(grid.col_widths_pt):
            col_letter = get_column_letter(i + 1)
            char_w = max(0.5, w_pt * _PT_TO_CHAR)
            # Clamp: tiny spacer columns stay tiny, content columns stay readable
            char_w = min(char_w, cfg.max_col_width)
            ws.column_dimensions[col_letter].width = char_w

        # ─────────── ROW HEIGHTS ─────────────────────────────────────────
        for i, h_pt in enumerate(grid.row_heights_pt):
            ws.row_dimensions[i + 1].height = max(4.0, h_pt * 0.95)

        # ─────────── RESOLVE COLLISIONS ──────────────────────────────────
        # If two GridCells map to overlapping merge ranges on the same row,
        # keep the first one and concatenate later ones.
        occupied: Dict[Tuple[int, int], GridCell] = {}   # (row, col) -> cell
        final_cells: List[GridCell] = []

        for gc in grid.cells:
            conflict = False
            for c in range(gc.start_col, gc.end_col + 1):
                if (gc.excel_row, c) in occupied:
                    # Append text to existing cell
                    occupied[(gc.excel_row, c)].text += " " + gc.text
                    conflict = True
                    break
            if not conflict:
                for c in range(gc.start_col, gc.end_col + 1):
                    occupied[(gc.excel_row, c)] = gc
                final_cells.append(gc)

        # ─────────── MERGED RANGES (pre-compute to avoid openpyxl conflicts) ──
        merges: List[Tuple[int, int, int, int]] = []  # (r, sc, r, ec)
        for gc in final_cells:
            if gc.end_col > gc.start_col:
                merges.append((gc.excel_row, gc.start_col, gc.excel_row, gc.end_col))

        # Apply merges first (openpyxl requires merge before styling)
        for r, sc, _, ec in merges:
            try:
                ws.merge_cells(
                    start_row=r, start_column=sc,
                    end_row=r, end_column=ec,
                )
            except Exception:
                pass  # already merged or overlap

        # ─────────── WRITE CELLS ─────────────────────────────────────────
        header_rows: Set[int] = set()
        total_rows: Set[int] = set()

        for gc in final_cells:
            if gc.is_header:
                header_rows.add(gc.excel_row)
            if gc.is_total:
                total_rows.add(gc.excel_row)

        for gc in final_cells:
            ws_cell = ws.cell(
                row=gc.excel_row,
                column=gc.start_col,
                value=gc.text,
            )

            # ── Font ──
            bold = gc.bold or gc.excel_row in header_rows or gc.excel_row in total_rows
            fs = gc.font_size if gc.font_size and gc.font_size > 0 else 10
            font_color = gc.text_color or "FF000000"
            ws_cell.font = Font(bold=bold, size=fs, color=font_color)

            # ── Alignment ──
            ws_cell.alignment = Alignment(
                horizontal=gc.align_h,
                vertical="center",
                wrap_text=True,
            )

            # ── Background fill ──
            if gc.bg_color:
                ws_cell.fill = PatternFill(fill_type="solid", fgColor=gc.bg_color)
            elif gc.excel_row in total_rows:
                ws_cell.fill = PatternFill(fill_type="solid", fgColor="FFFFF2CC")

        # ─────────── TABLE BORDERS ───────────────────────────────────────
        for (first_r, last_r, first_c, last_c) in grid.table_row_ranges:
            for r in range(first_r, last_r + 1):
                for c in range(first_c, last_c + 1):
                    cell = ws.cell(row=r, column=c)
                    cell.border = _THIN_BORDER

        # ─────────── PAGE SETUP ──────────────────────────────────────────
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True

        logger.debug(
            "Rendered page %d: %d cells, %d merges, %d cols x %d rows",
            grid.page_num, len(final_cells), len(merges),
            grid.total_cols, grid.total_rows,
        )
